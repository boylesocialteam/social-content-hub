#!/usr/bin/env python3
"""
Social Content Hub — automated build.

Reads the weekly Excel spreadsheets (promotions database, gaming plan) and the
Brandwatch schedule CSVs, and regenerates data.js in the shape the site expects.
This is the deterministic version of the transforms previously run by hand.

Run:
    python3 build/build_data.py            # build using build/config.json
    python3 build/build_data.py --check    # print to stdout, don't write

Rebuilds:   weeks[] (schedule + priorities), promoWeeks[], gamingWeeks[]
Preserves:  campaign (Calendar) and links — copied from the existing data.js.

Conventions (see Promo_Copy_Rulebook.md):
  * Copy is generated deterministically from the promotions database — same sheet,
    same copy. Editorial changes are made in the sheet (the control point).
  * Super Boost cards carry NO captions/platform link but KEEP the design box.
  * Emoji: 🚀 for boosts; 📲 on Facebook lines only; none elsewhere.
  * Tennis (Wimbledon) EPO is Monday-only; darts EPO posts on each match day.
"""
import json, re, os, sys, csv, datetime

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)


# ------------------------------------------------------------------ helpers ---
def load_config():
    with open(os.path.join(HERE, "config.json"), encoding="utf-8") as f:
        return json.load(f)


def src(cfg, name):
    return os.path.join(ROOT, cfg["sources_dir"], name)


def strip_time(s):
    return re.sub(r"\s*\([^)]*\)\s*$", "", str(s or "")).strip()


# ------------------------------------------------------- sport / link config ---
SPORT_META = {
    "Football":     ("World Cup",       "https://www.boylesports.com/promotions/online/football"),
    "Golf":         ("The Open",        "https://www.boylesports.com/sports/golf"),
    "Horse Racing": ("Horse Racing",    "https://www.boylesports.com/sports/horse-racing"),
    "Darts":        ("World Matchplay", "https://www.boylesports.com/sports/darts"),
    "Gaelic Games": ("GAA",             "https://www.boylesports.com/sports/gaelic-games"),
    "Tennis":       ("Wimbledon",       "https://www.boylesports.com/sports/tennis"),
}
FOOTBALL_LINK = SPORT_META["Football"][1]

CONTEXT = {  # safe, confirmed context only
    "usa": "co-hosts the USA", "canada": "co-hosts Canada", "mexico": "co-hosts Mexico",
    "argentina": "World Cup holders Argentina", "germany": "four-time winners Germany",
    "spain": "2010 winners Spain", "brazil": "five-time winners Brazil",
}
VERBS = ["take on", "face", "meet", "go head-to-head with", "lock horns with", "do battle with"]

COMP_SLUG = {"World Cup": "international-world-cup"}
TEAM_SLUG = {"bosnia": "bosnia-&-herzegovina", "dr congo": "dr-congo", "ivory coast": "ivory-coast",
             "curaçao": "curacao", "curacao": "curacao", "south korea": "south-korea", "usa": "usa"}


def team_ctx(name):
    return CONTEXT.get(name.strip().lower().replace("the ", ""), name.strip())


def art(name):
    """'the ' before an event name, unless it already starts with 'The'."""
    return "" if str(name).strip().lower().startswith("the ") else "the "


def teamslug(name):
    k = name.strip().lower()
    return TEAM_SLUG.get(k, re.sub(r"[^a-z0-9]+", "-", k).strip("-"))


def market_url(match):
    m = strip_time(match)
    if " v " not in m:
        return FOOTBALL_LINK
    a, b = [x.strip() for x in m.split(" v ", 1)]
    return f"https://www.boylesports.com/sports/football/event/international-world-cup/{teamslug(a)}-v-{teamslug(b)}"


# ---------------------------------------------------------- copy generators ---
def gen_copy(sport, ptype, match, pname, verb):
    """Return (twitter, facebook, platform_link). Empty strings where a channel is unused."""
    m = strip_time(match)
    label, page = SPORT_META.get(sport, (sport, ""))
    is_vs = " v " in m
    pt = ptype.lower()
    nums = re.findall(r"\d+", pname)

    # --- Bet Builder Boost (50 / 25) ---
    if "bbb" in pt or "bet builder boost" in pt:
        pct = "50" if "50" in ptype else "25"
        if is_vs:
            a, b = [x.strip() for x in m.split(" v ", 1)]
            body = f"Win {pct}% more with Bet Builder Boost as {team_ctx(a)} {verb} {b} in the {label}."
            url = market_url(m)
        else:
            body = f"Win {pct}% more with Bet Builder Boost on the {m}."
            url = page
        return f"🚀 {body}\n\nTap below to view the full market.", f"{body}\n\n📲 Full market — {url}", url

    # --- Super Boost: design box only (handled by caller; no copy) ---
    if "super boost" in pt:
        return "", "", ""

    # --- Bet & Get (incl. Multi / Squad / Accas / +5 if Qualify) ---
    if "bet & get" in pt or "bet and get" in pt:
        stake, free = (nums[0], nums[1]) if len(nums) >= 2 else (None, None)
        if "squad" in pname.lower():
            what = "on a Squad Builder"
        elif "multi" in pname.lower():
            what = "on a multiple"
        elif "acca" in pname.lower():
            what = "on Accas"
        else:
            what = "on a Bet Builder"
        loc = f"for {m}" + (f" in the {label}" if is_vs else "")
        if stake:
            body = f"Get a £/€{free} Free Bet when you bet £/€{stake} {what} {loc}."
            if "qualify" in pname.lower() and len(nums) >= 3:
                body = (f"Get a £/€{nums[1]} Free Bet when you bet £/€{nums[0]} {what} {loc} — "
                        f"and £/€{nums[2]} more if they qualify.")
        else:
            body = f"Bet & Get available {loc.replace('for ', 'on ')} — back a Bet Builder and claim your Free Bet."
        return f"{body}\n\nTap below to opt in.", f"{body}\n\n📲 Opt in — {page}", page

    # --- EPO (Early Payout) ---
    if pt == "epo":
        if sport == "Darts":
            body = ("Stay one step ahead of the drama with BOYLE Sports Early Payout!\n\n"
                    f"Get paid out when your player goes 4 legs ahead at the {m}, no matter how the match ends.")
            return f"{body}\n\nTap below to find out more.", f"{body}\n\n📲 Find out more — {page}", page
        # football / tennis EPO — Twitter only
        tw = (f"Get paid out before the final whistle with BoyleSports Early Payout at {label}! "
              f"When your team goes 2 goals ahead in our Match Betting market, we pay out.\n\nTap below to find out more.")
        return tw, "", page

    # --- Boost / Places (golf extra places) ---
    if "places" in pt or "places" in pname.lower():
        pl = re.search(r"(\d+)\s*Places", pname)
        n = pl.group(1) if pl else nums[0] if nums else "extra"
        ev = re.sub(r"\s*-\s*\d+\s*Places.*$", "", pname).strip() or m
        body = f"We're paying {n} PLACES on {art(ev)}{ev}."
        return f"{body}\n\nTap below to view the full market.", f"{body}\n\n📲 Full market — {page}", page

    # --- Outright Boost / Price Boost / PPB (Twitter only, 🚀) ---
    if "outright ppb" in pt or pt == "ppb":
        return f"🚀 Outright Price Boost on {art(m)}{m}!\n\nTap below to view the market.", "", page
    if "outright boost" in pt:
        return f"🚀 Boosted outright prices for {art(m)}{m}!\n\nTap below to view the market.", "", page
    if "price boost" in pt:
        return f"🚀 Price Boost on {art(m)}{m} — enhanced odds available now.\n\nTap below to add to your betslip.", "", page

    # --- Best Odds Guaranteed (racing, Twitter only) ---
    if "bog" in pt or "best odds" in pt:
        return ("Best Odds Guaranteed on all UK & Irish racing — take an early price and if the SP is "
                "bigger, we pay out at the bigger odds.\n\nTap below to view the racecard.", "", page)

    # --- fallback ---
    return "", "", page


# ---------------------------------------------------------- promotions page ---
DAYORD = {d: i for i, d in enumerate(
    ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday", "Sunday"])}


def build_promotions(cfg):
    import openpyxl
    wb = openpyxl.load_workbook(src(cfg, cfg["promotions_xlsx"]))
    ws = wb["Promotions"]
    hdr = {str(ws.cell(1, c).value).strip().lower(): c for c in range(1, ws.max_column + 1)}
    col = lambda n: hdr.get(n)
    items, meta = [], []
    verb_i = 0
    for r in range(2, ws.max_row + 1):
        day = ws.cell(r, col("day")).value
        if not day:
            continue
        # honour the "Include on Promotions Page?" flag — dedups promos that recur
        # across days so each shows only on its event day.
        inc_col = col("include on promotions page?")
        if inc_col and str(ws.cell(r, inc_col).value or "Y").strip().upper() == "N":
            continue
        sport = (ws.cell(r, col("sport")).value or "").strip()
        ptype = (ws.cell(r, col("promotion type")).value or "").strip()
        match = (ws.cell(r, col("event / match")).value or "").strip()
        pname = (ws.cell(r, col("promotion name")).value or "").strip()
        prio = (ws.cell(r, col("priority")).value or "").strip()
        note = (ws.cell(r, col("notes")).value or "").strip() if col("notes") else ""
        cell = ws.cell(r, col("asset link"))
        asset = cell.hyperlink.target if cell.hyperlink else ""
        if asset and not asset.startswith("http"):
            asset = ""  # skip local file:/// paths
        label = SPORT_META.get(sport, (sport, ""))[0]
        competition = label if sport != "Football" else "World Cup"

        # tennis EPO only on Monday
        if ptype.upper() == "EPO" and sport == "Tennis" and day.lower() != "monday":
            continue

        card = {"day": day, "match": strip_time(match), "competition": competition,
                "promo": pname, "type": ptype, "twitter": "", "facebook": "",
                "air": asset, "boylesports": "", "image": "", "note": note}

        if "super boost" not in ptype.lower():
            verb = VERBS[verb_i % len(VERBS)]; verb_i += 1
            tw, fb, link = gen_copy(sport, ptype, match, pname, verb)
            card.update(twitter=tw, facebook=fb, boylesports=link)
        items.append(card)
        meta.append({"match": strip_time(match), "type": ptype, "prio": prio, "day": day})
    return items, meta


def build_priorities(meta):
    rank = {"high": 0, "medium": 1, "low": 2, "": 3}
    groups = {}
    for m in meta:
        g = groups.setdefault(m["match"], {"types": [], "prio": 3, "day": 9})
        if m["type"] not in g["types"]:
            g["types"].append(m["type"])
        g["prio"] = min(g["prio"], rank.get(m["prio"].lower(), 3))
        g["day"] = min(g["day"], DAYORD.get(m["day"], 9))
    ordered = sorted(groups.items(), key=lambda kv: (kv[1]["prio"], -len(kv[1]["types"]), kv[1]["day"]))
    out = []
    for match, g in ordered[:5]:
        out.append({"title": match, "note": " · ".join(g["types"]),
                    "level": "high" if g["prio"] == 0 else "med"})
    return out


# --------------------------------------------------------------- schedule ------
PROMO_KW = ["bbb", "super boost", "price boost", "outright boost", "outright ppb", "ppb",
            "bet & get", "bet and get", "bet 10 get", "bet 5 get", "multi bet", "best odds",
            "12 places", "places", "epo", "money back", "bog", "boost - "]
SCHED_KW = ["schedule", "preview", "build-up", "build up", "tee times", "outrights", "outright"]


def classify(text):
    low = text.lower()
    if low.strip() == "lineups" or "lineups" in low:
        return "lineups"
    if any(k in low for k in PROMO_KW):
        return "promo"
    if any(k in low for k in SCHED_KW):
        return "schedule"
    return "fixture"


def build_schedule(cfg):
    days = {}
    for name in cfg.get("brandwatch_csvs", []):
        p = src(cfg, name)
        if not os.path.exists(p):
            continue
        with open(p, encoding="utf-8-sig") as f:
            for row in csv.reader(f):
                if len(row) < 2 or not row[0].strip():
                    continue
                dt = datetime.datetime.strptime(row[0].strip(), "%d/%m/%Y %H:%M")
                text = row[1].strip()
                d = days.setdefault(dt.date(), [])
                d.append({"time": dt.strftime("%H:%M"), "text": text, "cat": classify(text)})
    out = []
    for date in sorted(days):
        out.append({"date": date.isoformat(),
                    "label": date.strftime("%A %-d %b"),
                    "posts": sorted(days[date], key=lambda p: p["time"])})
    return out


# --------------------------------------------------------------- gaming page ---
def tc(s):
    s = str(s or "").strip()
    return s.title() if s.isupper() else s


_EMO = r"[\U0001F000-\U0001FAFF☀-➿⬀-⯿]"


def collapse_emoji(t):
    """Collapse doubled emoji flourishes (🤩🤩 / ✨ ✨) to a single, and ensure a
    leading emoji is followed by a space. Keeps mid-post emoji (gaming rulebook)."""
    t = (t or "").strip()
    t = re.sub(rf"({_EMO})(️?)(?:\s*\1\2?)+", r"\1\2", t)      # 🤩🤩 -> 🤩
    t = re.sub(rf"^({_EMO}️?)(?=\w)", r"\1 ", t)               # 🤩New -> 🤩 New
    return t


def to_ig(meta):
    return collapse_emoji(meta) + "\n\n📲 Click the link in our bio to learn more."


def link(base, btag):
    if not base:
        return ""
    u = base if str(base).endswith("/") else str(base) + "/"
    return u + (f"?btag={btag}" if btag else "")


def build_gaming(cfg, lo, hi):
    import openpyxl
    wb = openpyxl.load_workbook(src(cfg, cfg["gaming_xlsx"]))
    cards = []

    def inwin(d):
        if not isinstance(d, (datetime.datetime, datetime.date)):
            return False
        dd = d.date() if isinstance(d, datetime.datetime) else d
        return lo <= dd <= hi

    def fmt(d):
        dd = d.date() if isinstance(d, datetime.datetime) else d
        return dd.strftime("%-d %b %Y")

    for name in wb.sheetnames:
        low = name.lower()
        ws = wb[name]
        if "live casino" in low:
            for r in range(2, ws.max_row + 1):
                date = ws.cell(r, 4).value
                # real name is the promo (col 3) when present, else the post label (col 1)
                nm = ws.cell(r, 3).value or ws.cell(r, 1).value
                if not nm or not inwin(date):
                    continue
                meta = ws.cell(r, 10).value or ""
                bt = str(ws.cell(r, 13).value or "")
                mb = re.search(r"Meta btag[:\s]+(\S+)", bt)
                tb = re.search(r"Twitter btag[:\s]+(\S+)", bt)
                lk = ws.cell(r, 11).value
                cards.append({"section": "Live Casino", "name": tc(nm), "type": tc(ws.cell(r, 2).value),
                              "date": fmt(date), "meta": meta, "twitter": collapse_emoji(meta),
                              "instagram": to_ig(meta),
                              "design": ws.cell(r, 1).hyperlink.target if ws.cell(r, 1).hyperlink else "",
                              "facebook_link": link(lk, mb.group(1) if mb else ""),
                              "twitter_link": link(lk, tb.group(1) if tb else ""), "files": ""})
        elif "social media posts" in low:
            for r in range(3, ws.max_row + 1):
                nm, date = ws.cell(r, 1).value, ws.cell(r, 4).value  # game name is col 1
                if not nm or not inwin(date):
                    continue
                meta = ws.cell(r, 6).value or ""
                lk = ws.cell(r, 7).value
                cards.append({"section": "Gaming", "name": tc(nm), "type": tc(ws.cell(r, 2).value),
                              "date": fmt(date), "meta": meta, "twitter": collapse_emoji(meta),
                              "instagram": to_ig(meta),
                              "design": ws.cell(r, 1).hyperlink.target if ws.cell(r, 1).hyperlink else "",
                              "facebook_link": link(lk, ws.cell(r, 8).value),
                              "twitter_link": link(lk, ws.cell(r, 9).value), "files": ""})
    cards.sort(key=lambda c: datetime.datetime.strptime(c["date"], "%d %b %Y"))
    return cards


# --------------------------------------------------------------- assembly ------
def load_existing(cfg):
    p = os.path.join(ROOT, cfg["output"])
    if not os.path.exists(p):
        return {}
    raw = open(p, encoding="utf-8").read()
    m = re.search(r"window\.SCH_DATA\s*=\s*(\{[\s\S]*\});\s*$", raw.strip())
    return json.loads(m.group(1)) if m else {}


def write_data(cfg, data):
    with open(os.path.join(ROOT, cfg["output"]), "w", encoding="utf-8") as f:
        f.write("// Social Content Hub — weekly data. GENERATED by build/build_data.py. Do not hand-edit.\n")
        f.write("window.SCH_DATA = ")
        json.dump(data, f, ensure_ascii=False, indent=2)
        f.write(";\n")


def main():
    cfg = load_config()
    check = "--check" in sys.argv
    existing = load_existing(cfg)

    promos, meta = build_promotions(cfg)
    lo = datetime.date.fromisoformat(cfg["week_commencing"])
    hi = lo + datetime.timedelta(days=6)
    gaming = build_gaming(cfg, lo, hi)
    schedule = build_schedule(cfg)
    priorities = build_priorities(meta)

    label = cfg["week_label"]
    week = {"label": label, "commencing": cfg["week_commencing"], "theme": cfg["week_theme"],
            "updated": datetime.date.today().isoformat(), "priorities": priorities, "schedule": schedule}

    data = {"weeks": [week], "links": existing.get("links", []),
            "promoWeeks": [{"label": label, "items": promos}],
            "gamingWeeks": [{"label": label, "items": gaming}],
            "campaign": existing.get("campaign", {})}

    print(f"promos: {len(promos)} | gaming: {len(gaming)} | schedule days: {len(schedule)} | "
          f"priorities: {len(priorities)}", file=sys.stderr)
    if check:
        json.dump(data, sys.stdout, ensure_ascii=False, indent=2)
    else:
        write_data(cfg, data)
        print(f"✓ wrote {cfg['output']}", file=sys.stderr)


if __name__ == "__main__":
    main()
